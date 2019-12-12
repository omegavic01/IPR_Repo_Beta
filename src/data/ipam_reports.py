#!/usr/bin/python
"""
Copyright 2007 HVictor
Licensed to PSF under a Contributor Agreement.

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

     http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or
implied. See the License for the specific language governing
permissions and limitations under the License.
"""
from datetime import datetime
import time
import shutil
import os
import openpyxl
from openpyxl.styles import Alignment
from builder import DirectoryValues
from builder import DataFileNames


class IpamReports:
    """Takes the processed files and build the reports for IPR."""
    def __init__(self):
        self.data_filename_cls = DataFileNames()
        self.dir_cls = DirectoryValues()
        self.process_dir = self.dir_cls.processed_dir()
        self.reports_dir = self.dir_cls.reports_dir()
        self.ipam_filename = self.data_filename_cls.processed_filename()
        self.ipam_to_ipr_xlsx = os.path.join(self.process_dir,
                                             self.ipam_filename
                                             )
        self.date = self._get_file_date(self.ipam_to_ipr_xlsx
                                        )

    @staticmethod
    def _get_file_date(file):
        """Returns creation date of file."""
        date_str = time.ctime(os.path.getmtime(file)
                              )
        datetime_object = datetime.strptime(date_str, '%a %b %d %H:%M:%S %Y')
        return datetime_object.strftime('%Y%m%d')

    @staticmethod
    def _copy_data_over(source, template, final, sheet_name):
        """
        Reference to the site that was used for the below for loop:

        URL:
        https://stackoverflow.com/questions/44593705/how-to-copy-over-an-excel-
        sheet-to-another-workbook-in-python
        """
        source_wb = openpyxl.load_workbook(filename=source)
        source_ws = source_wb[sheet_name]
        try:
            template_wb = openpyxl.load_workbook(filename=template)
        except FileNotFoundError:
            print('Report by percent-BLANK.xlsx template not found.')
            return
        template_ws = template_wb.worksheets[1]
        for row in source_ws:
            for cell in row:
                template_ws[cell.coordinate].value = cell.value
        max_row = template_ws.max_row
        for row in template_ws.iter_rows(min_row=2, max_row=max_row,
                                         min_col=24, max_col=25):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
        template_wb.save(final)

    @staticmethod
    def _create_new_ipam_file_name_with_date_added(date, file_name):
        """Will work on xlsx extension files only."""
        date += '.xlsx'
        return file_name[:-5] + '-' + date

    @staticmethod
    def _create_new_percent_file_name_with_date_added(date, file_name):
        """Will work on xlsx extension files only."""
        date += '.xlsx'
        return file_name[:-10] + date

    @staticmethod
    def _create_ipam_report(processed_file, report_file):
        shutil.copy(processed_file, report_file)

    def generate_ipam_to_ipr_report(self):
        """Generates IPAM-to-IPR-(date).xlsx report for IPR."""
        ipam_report_with_date_filename = \
            self._create_new_ipam_file_name_with_date_added(
                self.date, self.ipam_filename)
        reports_ipam_to_ipr_xlsx = os.path.join(self.reports_dir,
                                                ipam_report_with_date_filename)
        self._create_ipam_report(self.ipam_to_ipr_xlsx,
                                 reports_ipam_to_ipr_xlsx)

    def generate_percent_report(self):
        """Generates IPR_Percent report xlsx file for IPR."""
#        percent_blank_filename = \
#            self.data_filename_cls.percent_blank_filename()
        percent_blank_xlsx = os.path.join(
            self.process_dir,
            self.data_filename_cls.percent_blank_filename()
        )
        percent_report_with_date_filename = \
            self._create_new_percent_file_name_with_date_added(
                self.date,
                self.data_filename_cls.percent_blank_filename()
            )
        percent_report_with_date_filename_and_path = os.path.join(
            self.reports_dir,
            percent_report_with_date_filename
        )
        self._copy_data_over(self.ipam_to_ipr_xlsx,
                             percent_blank_xlsx,
                             percent_report_with_date_filename_and_path,
                             'Summary')

    def generate_forecast_percent_report(self):
        """Generates IPR_Forecast_Percent report xlsx file for IPR."""
        percent_blank_forecast_filename = \
            self.data_filename_cls.percent_blank_forecast_filename()
        percent_blank_xlsx = os.path.join(self.process_dir,
                                          percent_blank_forecast_filename)
        percent_forecast_report_with_date_filename = \
            self._create_new_percent_file_name_with_date_added(
                self.date, percent_blank_forecast_filename)
        percent_forecast_report_with_date_filename_and_path = os.path.join(
            self.reports_dir,
            percent_forecast_report_with_date_filename)
        self._copy_data_over(
            self.ipam_to_ipr_xlsx,
            percent_blank_xlsx,
            percent_forecast_report_with_date_filename_and_path,
            'Summary_Forecast')
