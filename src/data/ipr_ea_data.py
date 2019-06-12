import os
import json
import pickle
import openpyxl
import logging
import requests
from dotenv import find_dotenv, load_dotenv


def _wr_out_ea_data(output_file, ea_data):
    ea_wb = openpyxl.Workbook()
    ea_sheet = ea_wb.create_sheet('EA Data')

    def _wr_ea_data(ea_data, col_id):
        ea_sheet.cell(row=1, column=col_id, value=ea_data['name'])
        for index, item in enumerate(ea_data['list_values']):
            mycell = ea_sheet.cell(row=index + 2, column=col_id)
            mycell.value = item['value']
    col_id = 0
    for i in ea_data:
        if i['type'] == 'ENUM':
            col_id += 1
            _wr_ea_data(i, col_id)
    if 'Sheet' in ea_wb.sheetnames:
        std = ea_wb['Sheet']
        ea_wb.remove(std)
    ea_wb.save(output_file)


def main():
    """
    Doc:
    Process:
    Output Files:
    """
    logger = logging.getLogger('ipr_ea_data.py')
    logger.info('Beginning of Script')
    logger.info('Building Paths and File names')

    # Build Directories
    raw_data_path = os.path.join(PROJECT_DIR, 'data', 'raw')
    reports_data_path = os.path.join(PROJECT_DIR, 'reports')

    # Build File and File path.
    ea_data_file = os.path.join(raw_data_path, 'ipr_ea_data.pkl')
    ea_report_file = os.path.join(reports_data_path, 'ipr_ea_data.xlsx')

    try:
        reatt = requests.get(PAYLOAD['url'] + "extensibleattributedef?"
                                              "_return_fields=list_values,"
                                              "comment,name,type",
                             auth=(PAYLOAD['username'], PAYLOAD['password']),
                             verify=False)
        reatt.raise_for_status()
    except requests.exceptions.ConnectionError as eaerrt:
        print("Can't reach Infoblox!  Check your VPN or Local access", eaerrt)
        exit()
    except requests.exceptions.HTTPError as eahrrt:
        print('Check your credentials!', eahrrt)
        exit()

    rutfeatt = reatt.content.decode('utf-8')
    rjsoneatt = json.loads(rutfeatt)
    with open(ea_data_file, 'wb') as fo:
        pickle.dump(rjsoneatt, fo)

    _wr_out_ea_data(ea_report_file, rjsoneatt)


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    # PAYLOAD for login to IPAM
    PAYLOAD = {
        'url': os.environ.get("DDI_URL"),
        'username': os.environ.get("DDI_USERNAME"),
        'password': os.environ.get("DDI_PASSWORD")
    }

    main()