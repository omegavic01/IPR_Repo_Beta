#!/usr/bin/python
import os
import openpyxl
from pandas import pandas as pd
from ipam_base_processing import BaseIpamProcessing

"""
The objective of this script is to update the Global Addresses for
IPR <data>.xlsx with a two digit country code in column F.

The Abbreviations.xlsx file needs to be in the data\raw directory.  Source
file is located in Teams-Internal.
"""


class GlobalAddress(BaseIpamProcessing):

    def __init__(self):
        BaseIpamProcessing.__init__(self)

        # Defined File Names
        self.country_abbreviations_file = os.path.join(
            self.dir_cls.raw_dir(),
            self.filename_cls.src_country_abbreviations()
        )

        self.global_street_addr_file = os.path.join(
            self.dir_cls.raw_dir(),
            self.filename_cls.src_global_street_addresses()
        )

    def return_country_abbreviations_file(self):
        return self.country_abbreviations_file

    def return_global_street_addr_file(self):
        return self.global_street_addr_file

    @staticmethod
    def return_global_street_addr_sheet():
        return 'Sheet1'

    @staticmethod
    def return_country_code_col():
        return 'Country Code'


class SourceDataValidation(GlobalAddress):

    def __init__(self):
        GlobalAddress.__init__(self)

    @staticmethod
    def check_for_file_existence(file):
        if os.path.isfile(file):
            return
        else:
            print(f"File does not exist: {file}")
            exit()

    def check_for_blank_country_codes(self):
        address_df = pd.read_excel(
            self.return_global_street_addr_file(),
            sheet_name=self.return_global_street_addr_sheet())
        address_df = address_df.loc[address_df[
            self.return_country_code_col()].isnull(), :]
        if address_df.empty:
            pass
        else:
            print(f'Could not find a country code for: '
                  f'{address_df["Country"]}')
            exit()


class CountryCodeUpdate(SourceDataValidation):

    def __init__(self):
        SourceDataValidation.__init__(self)

    def run_country_codes(self):
        abbreviation_data = self.compile_abbreviation_data()
        self.identify_and_write_output(abbreviation_data)

    def identify_and_write_output(self, abbr_dict):
        global_address_wb = openpyxl.load_workbook(
            self.return_global_street_addr_file())
        global_address_sheet = global_address_wb['Sheet1']

        # Compares global address list to country codes then updates col.
        for enum, row in enumerate(global_address_sheet.iter_rows()):
            if row[3].value in abbr_dict:
                global_address_sheet.cell(
                    enum + 1, 6).value = abbr_dict[row[3].value]
        global_address_wb.save(self.return_global_street_addr_file())

    def compile_abbreviation_data(self):
        abbreviation_wb = openpyxl.load_workbook(
            self.return_country_abbreviations_file())
        abbreviation_sheet = abbreviation_wb['Country and State']

        # Builds a dictionary of Country : 2x Digit Country code
        abbreviation_data_dict = {}
        for enum, row in enumerate(abbreviation_sheet.iter_rows()):
            if enum != 0:
                abbreviation_data_dict.update({row[0].value: row[1].value})
            else:
                continue

        abbreviation_wb.close()

        return abbreviation_data_dict


#sourcedatacheck = SourceDataValidation()
#sourcedatacheck.check_for_file_existence(
#    sourcedatacheck.return_country_abbreviations_file()
#)
#sourcedatacheck.check_for_file_existence(
#    sourcedatacheck.return_global_street_addr_file()
#)
country_music = CountryCodeUpdate()
country_music.run_country_codes()
country_music.check_for_blank_country_codes()
