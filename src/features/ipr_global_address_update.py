"""
The objective of this script is to update the Global Addresses for
IPR <data>.xlsx with a two digit country code in column F.

The Abbreviations.xlsx file needs to be in the data\raw directory.  Source
file is located in Teams-Internal.
"""
import os
import logging
import pickle
from xlrd import open_workbook
import openpyxl
from dotenv import find_dotenv, load_dotenv


def identify_and_write_output(abbr_dict, file):
    global_wb = openpyxl.load_workbook(filename=file)
    global_sheet = global_wb['Sheet1']
    for enum, row in enumerate(global_sheet.iter_rows()):
        templist = [cell.value for cell in row]
        if enum == 0:
            continue
        if templist[3] in abbr_dict:
            global_sheet.cell(enum + 1, 6).value = abbr_dict[templist[3]]
    global_wb.save(file)


def intake_data(abbr_file):
    """
    Loads source data (.xlsx) and .pkl's the data.
    :param: abbr_file - abbreviations data file path with filename
    :param: global_file - global address data file path with filename
    :return: abbr_data, global_data
    """
    abbr_wb = open_workbook(abbr_file)
    abbr_sheet = abbr_wb.sheet_by_index(0)

    # Builds a dictionary of Country : 2x Digit Country code
    abbr_data = {abbr_sheet.row_values(enum)[0]: abbr_sheet.row_values(enum)[1]
                 for enum in range(abbr_sheet.nrows)}
    return abbr_data


def main():
    """
    Controller for the functions configured.

    Two parameters that need to be looked at within this function are:
    Args:
    :manual object ipr_data_filename - needs updating based on need.
    :manual object addr_data_filename - needs updating based on need.
    """
    logger = logging.getLogger('ipr_global_address_update.py')
    logger.info('Beginning of Script')

    # Builds Directories to be used.
    raw_data_path = os.path.join(PROJECT_DIR, 'data', 'raw')
    processed_data_path = os.path.join(PROJECT_DIR, 'data', 'processed')

    # Update IPR and UNO Data sources.
    abbreviation_data_filename = 'Abbreviations.xlsx'
    global_addr_data_filename = 'Global Addresses for IPR 2019-05-01.xlsx'

    def compile_data():
        # Join file names to path's.
        abbr_data_file = os.path.join(raw_data_path,
                                      abbreviation_data_filename)
        global_data_file = os.path.join(raw_data_path,
                                        global_addr_data_filename)
        # Call's "intake_data" function.
        return intake_data(abbr_data_file)

    # Intake data and pickle.  Set to True for new data.  Else change to False.
    logger.info('Compiling data.')
    abbr_data = compile_data()
    logger.info('Compiling data: Completed')

    # For this function we are going to write the output to the file as it
    # loops through the data.
    identify_and_write_output(abbr_data,
                              os.path.join(raw_data_path,
                                           global_addr_data_filename))


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    main()
